from models.AlumnoModel import AlumnoModel
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import date
import os

class AlumnoController:

    def __init__(self):
        self.model = AlumnoModel()

    def obtener_alumnos(self):
        return self.model.listar_alumnos_activos()

    def obtener_calificaciones_alumno(self, id_alumno):
        return self.model.obtener_calificaciones_alumno(id_alumno)

    def guardar_alumno(self, nombre, apellido_paterno, apellido_materno, matricula, grupo, semestre, especialidad):
        if not nombre:
            return False, "El nombre es obligatorio"
        if not matricula:
            return False, "La matrícula es obligatoria"
        self.model.crear_alumno(nombre, apellido_paterno, apellido_materno, matricula, grupo, semestre, especialidad)
        return True, "Alumno registrado"

    def importar_excel(self, archivo, grado=None, grupo=None, especialidad=None, turno=None):
        if grado is None:
            grado = "2"
        if grupo is None:
            grupo = "D"
        if especialidad is None:
            especialidad = "Programación"
        if turno is None:
            turno = "Matutino"

        from datetime import datetime
        año_actual = datetime.now().strftime("%y")

        df = pd.read_excel(archivo, sheet_name=0, header=None)
        
        # Buscar la fila donde está "ALUMNO"
        alumno_row_idx = None
        alumno_col_idx = None
        
        for idx, row in df.iterrows():
            for col_idx, val in enumerate(row):
                if val is not None and str(val).upper().strip() == "ALUMNO":
                    alumno_row_idx = idx
                    alumno_col_idx = col_idx
                    break
            if alumno_row_idx is not None:
                break
        
        if alumno_row_idx is None:
            raise Exception("No se encontró la columna 'ALUMNO'")
        
        # Encontrar columnas de calificaciones (U1, U2, U3)
        calif_cols = {}
        for col_idx, val in enumerate(df.iloc[alumno_row_idx]):
            if val is not None:
                val_str = str(val).upper().strip()
                if val_str == "U1":
                    calif_cols[1] = col_idx
                elif val_str == "U2":
                    calif_cols[2] = col_idx
                elif val_str == "U3":
                    calif_cols[3] = col_idx
        
        # Extraer datos de alumnos
        alumnos_data = []
        for idx in range(alumno_row_idx + 1, len(df)):
            nombre = df.iloc[idx, alumno_col_idx]
            if pd.isna(nombre):
                continue
            
            nombre_str = str(nombre).strip()
            if not nombre_str or nombre_str.startswith("="):
                continue
            
            nombre_str = nombre_str.replace("PROTOTIPOS", "").replace("DUAL", "").strip()
            
            if len(nombre_str) < 3:
                continue
            
            # Buscar número de control existente
            numero_control = None
            for col_idx, val in enumerate(df.iloc[alumno_row_idx]):
                if val is not None and str(val).upper() in ["#CONTROL", "CONTROL", "MATRICULA", "NO"]:
                    control_val = df.iloc[idx, col_idx]
                    if pd.notna(control_val):
                        numero_control = str(int(control_val)) if isinstance(control_val, float) else str(control_val)
                    break
            
            # Si no hay número de control, generar uno
            if not numero_control:
                consecutivo = idx + 1
                nivel = "3"
                plantel = "080606"
                tipo = "1"
                numero = str(consecutivo).zfill(4)
                numero_control = f"{año_actual}{nivel}{plantel}{tipo}{numero}"
            
            # Extraer calificaciones
            calificaciones = {}
            for parcial, col_idx in calif_cols.items():
                try:
                    valor = df.iloc[idx, col_idx]
                    if pd.notna(valor):
                        calificaciones[parcial] = min(float(valor), 10)
                except:
                    pass
            
            alumnos_data.append({
                "nombre": nombre_str,
                "numero_control": numero_control,
                "calificaciones": calificaciones
            })
        
        if not alumnos_data:
            raise Exception("No se encontraron alumnos en el archivo")
        
        id_materia = 2
        contador = 0
        
        for data in alumnos_data:
            nombre_completo = data["nombre"]
            numero_control = data["numero_control"]
            calificaciones = data["calificaciones"]
            
            partes = nombre_completo.split()
            
            if len(partes) >= 3:
                apellido_paterno = partes[0]
                apellido_materno = partes[1]
                nombre = " ".join(partes[2:])
            elif len(partes) == 2:
                apellido_paterno = partes[0]
                apellido_materno = ""
                nombre = partes[1]
            else:
                apellido_paterno = ""
                apellido_materno = ""
                nombre = partes[0]
            
            nombre = nombre.replace("(Atencion)", "").strip()
            
            if not self.model.existe_matricula(numero_control):
                self.model.crear_alumno(
                    nombre, apellido_paterno, apellido_materno,
                    numero_control, grupo, str(grado), especialidad
                )
            else:
                id_alumno = self.model.obtener_id_por_matricula(numero_control)
                self.model.actualizar_alumno(
                    id_alumno, nombre, apellido_paterno, apellido_materno,
                    numero_control, grupo, str(grado), especialidad
                )
            
            id_alumno = self.model.obtener_id_por_matricula(numero_control)
            
            for parcial, calificacion in calificaciones.items():
                self.model.crear_calificacion(id_alumno, id_materia, parcial, calificacion)
            
            contador += 1
        
        return contador

    def generar_plantilla(self, grupo, semestre, especialidad):
        directorio_actual = os.path.dirname(__file__)
        directorio_src = os.path.dirname(directorio_actual)
        directorio_raiz = os.path.dirname(directorio_src)
        ruta_archivo = os.path.join(directorio_raiz, "Plantilla de grupo.xlsx")
        return "file:///" + ruta_archivo.replace("\\", "/")

    def exportar_excel(self):
        alumnos = self.model.listar_alumnos()
        wb = Workbook()
        ws = wb.active
        ws.title = "Alumnos"
        encabezados = ["ID", "Nombre", "Apellido Paterno", "Apellido Materno", "Matrícula", "Grupo", "Semestre", "Especialidad", "Estatus"]
        for col, encabezado in enumerate(encabezados, start=1):
            ws.cell(row=1, column=col, value=encabezado)
        for row_idx, alumno in enumerate(alumnos, start=2):
            ws.cell(row=row_idx, column=1, value=alumno.get("id_alumno", ""))
            ws.cell(row=row_idx, column=2, value=alumno.get("nombre", ""))
            ws.cell(row=row_idx, column=3, value=alumno.get("apellido_paterno", ""))
            ws.cell(row=row_idx, column=4, value=alumno.get("apellido_materno", ""))
            ws.cell(row=row_idx, column=5, value=alumno.get("matricula", ""))
            ws.cell(row=row_idx, column=6, value=alumno.get("grupo", ""))
            ws.cell(row=row_idx, column=7, value=alumno.get("semestre", ""))
            ws.cell(row=row_idx, column=8, value=alumno.get("especialidad", ""))
            ws.cell(row=row_idx, column=9, value=alumno.get("estatus", "Activo"))
        archivo = "Alumnos_exportados.xlsx"
        wb.save(archivo)
        return archivo

    def obtener_asistencias_alumno(self, id_alumno):
        return self.model.obtener_asistencias_alumno(id_alumno)

    def crear_asistencia(self, id_alumno, fecha, estado):
        if hasattr(fecha, 'date'):
            fecha = fecha.date()
        self.model.crear_asistencia(id_alumno, fecha, estado)
        return True, "Asistencia registrada correctamente"

    def corregir_nombres_existentes(self):
        alumnos = self.model.listar_alumnos()
        corregidos = 0
        
        for alumno in alumnos:
            nombre_actual = alumno.get('nombre', '')
            ap_paterno_actual = alumno.get('apellido_paterno', '')
            ap_materno_actual = alumno.get('apellido_materno', '')
            
            if len(nombre_actual) < 10 and len(ap_paterno_actual) < 10:
                nombre_completo = f"{nombre_actual} {ap_paterno_actual} {ap_materno_actual}"
                partes = nombre_completo.split()
                
                if len(partes) >= 3:
                    nuevo_apellido_paterno = partes[0]
                    nuevo_apellido_materno = partes[1]
                    nuevo_nombre = " ".join(partes[2:])
                    
                    self.model.actualizar_alumno(
                        alumno['id_alumno'],
                        nuevo_nombre,
                        nuevo_apellido_paterno,
                        nuevo_apellido_materno,
                        alumno['matricula'],
                        alumno['grupo'],
                        alumno['semestre'],
                        alumno['especialidad']
                    )
                    corregidos += 1
        
        return corregidos
    def dar_de_baja_alumno(self, id_alumno):
        """Cambia el estatus del alumno a 'Baja'"""
        self.model.dar_de_baja_alumno(id_alumno)
        return True, "Alumno dado de baja correctamente"

    def eliminar_alumno(self, id_alumno):
        """Eliminación física permanente"""
        self.model.eliminar_alumno(id_alumno)
        return True, "Alumno eliminado permanentemente"